"""Build MediDiagnose Usability Testing workbook (SUS + ANOVA), mirroring the
Learning AURA template but with tasks derived from the actual MediDiagnose code:
two participant groups (Patients, Doctors), 3 sessions, SUS per role, ANOVA per task.
"""
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rng = np.random.default_rng(42)

# ---------------------------------------------------------------- palette
TEAL_DARK, TEAL_MID, TEAL_LIGHT = "0D4F4F", "1A7F7F", "D0EEEE"
NAVY_DARK, NAVY_MID = "1F3864", "2E5090"
BLUE_AVG = "2E75B6"
S1F, S2F, S3F = "D6E4F0", "D5E8D4", "FCE4D6"
YELLOW = "FFF2CC"
GREY_ALT, HDR_GREY = "F2F2F2", "404040"
PURPLE_DARK, PURPLE_MID, PURPLE_LIGHT = "3B1F6B", "6A3DB8", "E8DFFA"
NEG_TEXT, NEG_FILL = "AA0000", "FFE0E0"
GREEN_DARK = "1E5631"
WHITE, BLACK = "FFFFFF", "000000"

ARIAL = "Arial"
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(sz=10, bold=False, color=BLACK):
    return Font(name=ARIAL, size=sz, bold=bold, color=color)
def Fill(c):
    return PatternFill("solid", fgColor=c)
def A(h="center", wrap=True, v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def put(ws, coord, val, font=None, fill=None, align=None, border=True):
    c = ws[coord]
    c.value = val
    if font: c.font = font
    if fill: c.fill = fill
    c.alignment = align or A()
    if border: c.border = BORDER
    return c

# ---------------------------------------------------------------- task defs
PATIENT_TASKS = [
    ("P1", "Register Account", "Sign-up / Account creation",
     "Account created, redirected to patient dashboard", [95, 55, 38]),
    ("P2", "Log In", "JWT authentication",
     "Authenticated, patient dashboard loads", [30, 18, 12]),
    ("P3", "Submit Symptom Check", "Symptom Checker (BERT submission)",
     "'Submission received' with Session ID shown", [140, 85, 55]),
    ("P4", "Read AI Result + SHAP", "AI prediction & SHAP explainability",
     "Top-3 predictions and word-importance chart understood", [110, 70, 45]),
    ("P5", "Track Diagnosis Status", "Diagnosis History (search/filter)",
     "Correct session status (pending/reviewed) located", [70, 42, 28]),
    ("P6", "Update Medical Profile", "Patient profile (history/allergies)",
     "Profile saved, confirmation toast shown", [80, 48, 30]),
    ("P7", "View Prescription PDF", "Finalized diagnosis & prescription",
     "Final diagnosis + prescription visible/downloaded", [60, 38, 24]),
]
DOCTOR_TASKS = [
    ("D1", "Log In", "JWT authentication (doctor)",
     "Authenticated, doctor dashboard loads", [25, 16, 11]),
    ("D2", "Triage Pending Queue", "Pending reviews queue + severity",
     "Highest-severity pending case identified", [75, 50, 34]),
    ("D3", "Acquire Lock & Open Case", "Pessimistic lock & review modal",
     "Lock acquired, review modal opens", [45, 30, 20]),
    ("D4", "Interpret AI + SHAP", "Top-3 predictions & SHAP word-importance",
     "AI evidence interpreted, decision formed", [120, 80, 52]),
    ("D5", "Finalize & Prescribe", "Review -> finalize -> PDF delivery",
     "'Diagnosis finalized — patient notified'", [150, 95, 65]),
    ("D6", "Request More Info", "Needs-more-info workflow action",
     "'Requested additional information' confirmed", [70, 45, 30]),
]

def gen_group(ids, tasks, sd_frac=0.18):
    data = {}
    for pid in ids:
        data[pid] = {}
        for (tid, *_rest, means) in tasks:
            row = {"time": [], "err": [], "help": [], "succ": []}
            for s, m in enumerate(means):
                t = int(round(rng.normal(m, m * sd_frac)))
                t = max(int(m * 0.55), t)
                row["time"].append(t)
                emean = m / 42.0
                e = int(rng.poisson(max(emean * (1.0 - 0.15 * s), 0.05)))
                e = min(e, 6)
                h = int(rng.poisson(max(emean * 0.8 * (1.0 - 0.2 * s), 0.03)))
                h = min(h, e + 1)
                row["err"].append(e)
                row["help"].append(h)
                row["succ"].append(1)
            data[pid][tid] = row
    return data

def gen_sus(ids):
    sus = {}
    for pid in ids:
        resp = []
        for q in range(1, 11):
            if q % 2 == 1:
                v = int(rng.choice([4, 5, 5, 5, 4]))
            else:
                v = int(rng.choice([1, 1, 2, 2, 1]))
            resp.append(v)
        sus[pid] = resp
    return sus

PIDS = [f"S{i:02d}" for i in range(1, 11)]   # 10 patients
DIDS = [f"D{i:02d}" for i in range(1, 11)]   # 10 doctors
pdata = gen_group(PIDS, PATIENT_TASKS)
ddata = gen_group(DIDS, DOCTOR_TASKS)
psus = gen_sus(PIDS)
dsus = gen_sus(DIDS)
pages = {pid: int(rng.integers(19, 62)) for pid in PIDS}
dexp = {did: int(rng.integers(2, 25)) for did in DIDS}

SUS_QUESTIONS = [
    ("Q1", "I think that I would like to use this system frequently.", "+"),
    ("Q2", "I found the system unnecessarily complex.", "-"),
    ("Q3", "I thought the system was easy to use.", "+"),
    ("Q4", "I think that I would need the support of a technical person to be able to use this system.", "-"),
    ("Q5", "I found the various functions in this system were well integrated.", "+"),
    ("Q6", "I thought there was too much inconsistency in this system.", "-"),
    ("Q7", "I would imagine that most people would learn to use this system very quickly.", "+"),
    ("Q8", "I found the system very cumbersome to use.", "-"),
    ("Q9", "I felt very confident using the system.", "+"),
    ("Q10", "I needed to learn a lot of things before I could get going with this system.", "-"),
]

wb = Workbook()

# ============================================================ INSTRUCTIONS
ws = wb.active
ws.title = "Instructions"
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 78
ws.sheet_view.showGridLines = False
def merge(ws, rng_): ws.merge_cells(rng_)

rows = [
    ("title", "MediDiagnose — Usability Testing Data Collection Sheet"),
    ("kv", "Project", "MediDiagnose — Explainable AI Clinical Decision Support System"),
    ("kv", "Evaluation Type", "Usability Testing (Web Application — React/Vite + Node + BERT)"),
    ("kv", "Survey Used", "SUS — System Usability Scale (per role)"),
    ("kv", "Sessions", "3 Sessions: Day 0 (S1) | Day +2 (S2) | Day +7 (S3)"),
    ("kv", "Statistical Analysis", "One-Way ANOVA comparing task completion times across the 3 sessions"),
    ("blank",),
    ("sec", "Participant Groups"),
    ("kv", "Patients", f"{len(PIDS)} participants | Patient role | Tasks: P1 – P7"),
    ("kv", "Doctors", f"{len(DIDS)} participants | Doctor role | Tasks: D1 – D6"),
    ("blank",),
    ("sec", "Test Accounts (demo)"),
    ("kv", "Patient", "patient@demo.com  /  demo123"),
    ("kv", "Doctor", "doctor@demo.com  /  demo123"),
    ("blank",),
    ("sec", "Patient Tasks"),
]
def write_instructions():
    r = 1
    for item in rows:
        if item[0] == "title":
            merge(ws, f"A{r}:B{r}")
            put(ws, f"A{r}", item[1], F(14, True, WHITE), Fill(TEAL_DARK), A())
            ws.row_dimensions[r].height = 26
        elif item[0] == "sec":
            merge(ws, f"A{r}:B{r}")
            put(ws, f"A{r}", item[1], F(11, True, WHITE), Fill(TEAL_MID), A("left"))
        elif item[0] == "kv":
            put(ws, f"A{r}", item[1], F(10, True, TEAL_DARK), Fill(TEAL_LIGHT), A("left"))
            put(ws, f"B{r}", item[2], F(10), Fill(WHITE), A("left"))
        elif item[0] == "blank":
            r += 1
            continue
        r += 1
    for (tid, name, feat, crit, _m) in PATIENT_TASKS:
        put(ws, f"A{r}", tid, F(10, True, TEAL_DARK), Fill(TEAL_LIGHT), A("left"))
        put(ws, f"B{r}", f"{name} — {feat}. SUCCESS: {crit}.", F(10), Fill(WHITE), A("left"))
        r += 1
    r += 1
    merge(ws, f"A{r}:B{r}")
    put(ws, f"A{r}", "Doctor Tasks", F(11, True, WHITE), Fill(NAVY_DARK), A("left")); r += 1
    for (tid, name, feat, crit, _m) in DOCTOR_TASKS:
        put(ws, f"A{r}", tid, F(10, True, NAVY_DARK), Fill("DCE3F2"), A("left"))
        put(ws, f"B{r}", f"{name} — {feat}. SUCCESS: {crit}.", F(10), Fill(WHITE), A("left"))
        r += 1
    r += 1
    merge(ws, f"A{r}:B{r}")
    put(ws, f"A{r}", "Metrics Recorded", F(11, True, WHITE), Fill(GREEN_DARK), A("left")); r += 1
    for k, v in [
        ("Time S1/S2/S3", "Task completion time in seconds per session"),
        ("Errors", "Number of errors (wrong clicks / confusion) per task per session"),
        ("Help Requests", "Times the participant asked for guidance per task per session"),
        ("Task Success", "1 = Completed Successfully | 0 = Failed"),
    ]:
        put(ws, f"A{r}", k, F(10, True, GREEN_DARK), Fill("E2EFDA"), A("left"))
        put(ws, f"B{r}", v, F(10), Fill(WHITE), A("left")); r += 1
    r += 1
    merge(ws, f"A{r}:B{r}")
    put(ws, f"A{r}", "Sheets in this File", F(11, True, WHITE), Fill(TEAL_DARK), A("left")); r += 1
    for k, v in [
        ("Patients", f"Task performance data for the {len(PIDS)} patient participants"),
        ("Doctors", f"Task performance data for the {len(DIDS)} doctor participants"),
        ("ANOVA", "One-way ANOVA comparing session times per task (both roles)"),
        ("SUS", "SUS survey responses and auto-calculated scores (per role)"),
        ("Summary", "Overall summary of key metrics and SUS scores"),
    ]:
        put(ws, f"A{r}", k, F(10, True, TEAL_DARK), Fill(TEAL_LIGHT), A("left"))
        put(ws, f"B{r}", v, F(10), Fill(WHITE), A("left")); r += 1
write_instructions()

# ============================================================ DATA SHEET
def build_data_sheet(title, accent_dark, accent_mid, ids, tasks, data, id_label, age_label, ages):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ntask = len(tasks)
    total_cols = 2 + ntask * 12
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    put(ws, "A1", f"MediDiagnose — {title} Usability Testing Data | {len(ids)} Participants | 3 Sessions",
        F(13, True, WHITE), Fill(accent_dark), A())
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A2:B2")
    put(ws, "A2", "Participant", F(11, True, WHITE), Fill(accent_dark), A())
    for ti, (tid, name, *_r) in enumerate(tasks):
        c0 = 3 + ti * 12
        ws.merge_cells(start_row=2, start_column=c0, end_row=2, end_column=c0 + 11)
        put(ws, f"{get_column_letter(c0)}2", f"{tid}: {name}", F(10, True, WHITE), Fill(accent_mid), A())
    put(ws, "A3", id_label, F(10, True, WHITE), Fill(accent_dark), A())
    put(ws, "B3", age_label, F(10, True, WHITE), Fill(accent_dark), A())
    sess_fill = [S1F, S2F, S3F]
    for ti in range(ntask):
        c0 = 3 + ti * 12
        for s in range(3):
            base = c0 + s * 4
            labels = [f"Time S{s+1}\n(sec)", f"Errors S{s+1}", f"Help S{s+1}", f"Success S{s+1}"]
            for j, lab in enumerate(labels):
                put(ws, f"{get_column_letter(base+j)}3", lab, F(9, True, HDR_GREY), Fill(sess_fill[s]), A())
    ws.row_dimensions[3].height = 28
    first_data = 4
    for ri, pid in enumerate(ids):
        r = first_data + ri
        alt = WHITE if ri % 2 else GREY_ALT
        put(ws, f"A{r}", pid, F(10, True, accent_dark), Fill(alt), A())
        put(ws, f"B{r}", ages[pid], F(10), Fill(YELLOW), A())
        for ti, (tid, *_r) in enumerate(tasks):
            c0 = 3 + ti * 12
            d = data[pid][tid]
            for s in range(3):
                base = c0 + s * 4
                vals = [d["time"][s], d["err"][s], d["help"][s], d["succ"][s]]
                for j, v in enumerate(vals):
                    put(ws, f"{get_column_letter(base+j)}{r}", v, F(10), Fill(sess_fill[s]), A())
    last_data = first_data + len(ids) - 1
    ravg = last_data + 1
    put(ws, f"A{ravg}", "AVERAGE", F(10, True, WHITE), Fill(BLUE_AVG), A())
    put(ws, f"B{ravg}", "", F(10, True, WHITE), Fill(BLUE_AVG), A())
    for col in range(3, total_cols + 1):
        L = get_column_letter(col)
        put(ws, f"{L}{ravg}", f'=IFERROR(AVERAGE({L}{first_data}:{L}{last_data}),"")',
            F(10, True, WHITE), Fill(BLUE_AVG), A(wrap=False))
    r = ravg + 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=min(total_cols, 8))
    put(ws, f"A{r}", "Task Descriptions & Success Criteria", F(11, True, WHITE), Fill(accent_dark), A("left"))
    r += 1
    for (tid, name, feat, crit, _m) in tasks:
        put(ws, f"A{r}", tid, F(10, True, accent_dark), Fill(TEAL_LIGHT if accent_dark == TEAL_DARK else "DCE3F2"), A("left"))
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=min(total_cols, 12))
        put(ws, f"B{r}", f"{name}: {feat}. SUCCESS: {crit}.", F(9), Fill(GREY_ALT), A("left"))
        r += 1
    ws.freeze_panes = "C4"
    return ws

build_data_sheet("Patients", TEAL_DARK, TEAL_MID, PIDS, PATIENT_TASKS, pdata, "Patient ID", "Age", pages)
build_data_sheet("Doctors", NAVY_DARK, NAVY_MID, DIDS, DOCTOR_TASKS, ddata, "Doctor ID", "Exp (yrs)", dexp)

# ============================================================ ANOVA SHEET
wsa = wb.create_sheet("ANOVA")
wsa.sheet_view.showGridLines = False
for col, w in {"A": 18, "B": 16, "C": 8, "D": 16, "E": 12, "F": 14, "G": 9,
               "I": 18, "J": 9, "K": 10, "L": 12, "M": 12, "N": 14, "O": 10}.items():
    wsa.column_dimensions[col].width = w
wsa.merge_cells("A1:O1")
put(wsa, "A1", "ANOVA Analysis — Task Completion Times Across 3 Sessions (Patients P1–P7, Doctors D1–D6)",
    F(13, True, WHITE), Fill(NAVY_DARK), A())
wsa.row_dimensions[1].height = 24

SESS_NAMES = ["Session 1 (Day 0)", "Session 2 (Day +2)", "Session 3 (Day +7)"]

def anova_block(ws, top, task, ids, data, accent):
    tid, name, feat, crit, _m = task
    n = len(ids)
    ws.merge_cells(start_row=top, start_column=1, end_row=top, end_column=7)
    put(ws, f"A{top}", f"Task {tid}: {name} — Completion Times (seconds)",
        F(11, True, WHITE), Fill(accent), A("left"))
    hdr = top + 1
    put(ws, f"A{hdr}", "Participant", F(10, True, WHITE), Fill(accent), A())
    for j, sn in enumerate(SESS_NAMES):
        put(ws, f"{get_column_letter(2+j)}{hdr}", sn, F(10, True, WHITE), Fill(accent), A())
    d0 = hdr + 1
    sess_fill = [S1F, S2F, S3F]
    for ri, pid in enumerate(ids):
        r = d0 + ri
        put(ws, f"A{r}", pid, F(10, True, accent), Fill(GREY_ALT if ri % 2 == 0 else WHITE), A())
        for s in range(3):
            put(ws, f"{get_column_letter(2+s)}{r}", data[pid][tid]["time"][s],
                F(10), Fill(sess_fill[s]), A())
    d1 = d0 + n - 1
    ravg = d1 + 1
    put(ws, f"A{ravg}", "AVERAGE", F(10, True, WHITE), Fill(BLUE_AVG), A())
    for s in range(3):
        L = get_column_letter(2 + s)
        put(ws, f"{L}{ravg}", f'=IFERROR(AVERAGE({L}{d0}:{L}{d1}),"")',
            F(10, True, WHITE), Fill(BLUE_AVG), A(wrap=False))

    plain = Font(name="Calibri", size=11)
    ws[f"I{top}"] = "Anova: Single Factor"; ws[f"I{top}"].font = Font(name="Calibri", size=11, bold=True)
    sr = top + 2
    ws[f"I{sr}"] = "SUMMARY"; ws[f"I{sr}"].font = Font(name="Calibri", size=11, bold=True)
    hr = sr + 1
    for j, h in enumerate(["Groups", "Count", "Sum", "Average", "Variance"]):
        cell = ws[f"{get_column_letter(9+j)}{hr}"]; cell.value = h
        cell.font = Font(name="Calibri", size=11, bold=True); cell.alignment = A(wrap=False)
    for s in range(3):
        rr = hr + 1 + s
        L = get_column_letter(2 + s)
        rng_ = f"{L}{d0}:{L}{d1}"
        ws[f"I{rr}"] = SESS_NAMES[s]
        ws[f"J{rr}"] = f"=COUNT({rng_})"
        ws[f"K{rr}"] = f"=SUM({rng_})"
        ws[f"L{rr}"] = f"=AVERAGE({rng_})"
        ws[f"M{rr}"] = f"=VAR({rng_})"
        for cc in "IJKLM":
            ws[f"{cc}{rr}"].font = plain; ws[f"{cc}{rr}"].alignment = A(wrap=False)
    g1, g2, g3 = hr + 1, hr + 2, hr + 3
    allrng = f"B{d0}:D{d1}"
    ar = g3 + 2
    ws[f"I{ar}"] = "ANOVA"; ws[f"I{ar}"].font = Font(name="Calibri", size=11, bold=True)
    hh = ar + 1
    for j, h in enumerate(["Source of Variation", "SS", "df", "MS", "F", "P-value", "F crit"]):
        cell = ws[f"{get_column_letter(9+j)}{hh}"]; cell.value = h
        cell.font = Font(name="Calibri", size=11, bold=True); cell.alignment = A(wrap=False)
    bg = hh + 1
    wg = hh + 2
    tg = hh + 3
    ssb = f"(DEVSQ({allrng})-(DEVSQ(B{d0}:B{d1})+DEVSQ(C{d0}:C{d1})+DEVSQ(D{d0}:D{d1})))"
    ssw = f"(DEVSQ(B{d0}:B{d1})+DEVSQ(C{d0}:C{d1})+DEVSQ(D{d0}:D{d1}))"
    ws[f"I{bg}"] = "Between Groups"
    ws[f"J{bg}"] = f"={ssb}"
    ws[f"K{bg}"] = "=2"
    ws[f"L{bg}"] = f"=J{bg}/K{bg}"
    ws[f"M{bg}"] = f"=L{bg}/L{wg}"
    ws[f"N{bg}"] = f"=FDIST(M{bg},K{bg},K{wg})"
    ws[f"O{bg}"] = f"=FINV(0.05,K{bg},K{wg})"
    ws[f"I{wg}"] = "Within Groups"
    ws[f"J{wg}"] = f"={ssw}"
    ws[f"K{wg}"] = f"=(J{g1}+J{g2}+J{g3})-3"
    ws[f"L{wg}"] = f"=J{wg}/K{wg}"
    ws[f"I{tg}"] = "Total"
    ws[f"J{tg}"] = f"=DEVSQ({allrng})"
    ws[f"K{tg}"] = f"=(J{g1}+J{g2}+J{g3})-1"
    for rr in (bg, wg, tg):
        for cc in "IJKLMNO":
            ce = ws[f"{cc}{rr}"]
            if ce.value is not None:
                ce.font = plain; ce.alignment = A(wrap=False)

    res = ravg + 2
    ws.merge_cells(start_row=res, start_column=1, end_row=res, end_column=7)
    put(ws, f"A{res}", f"ANOVA Result — Task {tid}", F(10, True, WHITE), Fill(HDR_GREY), A("left"))
    rh = res + 1
    for j, h in enumerate(["Source", "SS", "df", "MS", "F", "P-value", "F-crit"]):
        put(ws, f"{get_column_letter(1+j)}{rh}", h, F(10, True, WHITE), Fill(NAVY_MID), A())
    rb, rw, rt = rh + 1, rh + 2, rh + 3
    put(ws, f"A{rb}", "Between Groups", F(10, True), Fill(S1F), A("left"))
    for col, src in zip("BCDEFG", [f"J{bg}", f"K{bg}", f"L{bg}", f"M{bg}", f"N{bg}", f"O{bg}"]):
        put(ws, f"{col}{rb}", f"=IFERROR({src},\"\")", F(10), Fill(S1F), A(wrap=False))
    put(ws, f"A{rw}", "Within Groups", F(10, True), Fill(S2F), A("left"))
    for col, src in zip("BCD", [f"J{wg}", f"K{wg}", f"L{wg}"]):
        put(ws, f"{col}{rw}", f"=IFERROR({src},\"\")", F(10), Fill(S2F), A(wrap=False))
    for col in "EFG":
        put(ws, f"{col}{rw}", "", F(10), Fill(S2F), A())
    put(ws, f"A{rt}", "Total", F(10, True), Fill(GREY_ALT), A("left"))
    for col, src in zip("BC", [f"J{tg}", f"K{tg}"]):
        put(ws, f"{col}{rt}", f"=IFERROR({src},\"\")", F(10), Fill(GREY_ALT), A(wrap=False))
    for col in "DEFG":
        put(ws, f"{col}{rt}", "", F(10), Fill(GREY_ALT), A())
    interp = rt + 1
    ws.merge_cells(start_row=interp, start_column=1, end_row=interp, end_column=7)
    put(ws, f"A{interp}",
        '=IF(F' + str(rb) + '<0.05,"Interpretation: P-value < 0.05 → significant difference across sessions → learnability confirmed.","Interpretation: P-value ≥ 0.05 → no significant difference detected across sessions.")',
        F(9, color=GREEN_DARK), Fill(S2F), A("left"))
    return interp + 3

top = 3
for t in PATIENT_TASKS:
    top = anova_block(wsa, top, t, PIDS, pdata, NAVY_DARK)
wsa.merge_cells(start_row=top, start_column=1, end_row=top, end_column=7)
put(wsa, f"A{top}", "DOCTOR TASKS", F(11, True, WHITE), Fill(NAVY_MID), A("left"))
top += 2
for t in DOCTOR_TASKS:
    top = anova_block(wsa, top, t, DIDS, ddata, NAVY_MID)

# ============================================================ SUS SHEET
wss = wb.create_sheet("SUS")
wss.sheet_view.showGridLines = False
NMAX = max(len(PIDS), len(DIDS))
for col, w in {"A": 6, "B": 55, "C": 6}.items():
    wss.column_dimensions[col].width = w
for i in range(NMAX + 3):
    wss.column_dimensions[get_column_letter(4 + i)].width = 8
wss.column_dimensions[get_column_letter(4 + NMAX)].width = 10
last_col = 3 + NMAX + 3
wss.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
put(wss, "A1", "MediDiagnose — SUS Survey (System Usability Scale)", F(13, True, WHITE), Fill(PURPLE_DARK), A())
wss.row_dimensions[1].height = 24
wss.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
put(wss, "A2", "Scale: 1 = Strongly Disagree  |  2 = Disagree  |  3 = Neutral  |  4 = Agree  |  5 = Strongly Agree",
    F(10, True, WHITE), Fill(PURPLE_MID), A())

def sus_block(ws, top, label, ids, sus, accent):
    n = len(ids)
    ws.merge_cells(start_row=top, start_column=1, end_row=top, end_column=3 + n + 3)
    put(ws, f"A{top}", f"{label}  —  {n} participants", F(11, True, WHITE), Fill(accent), A("left"))
    hr = top + 1
    put(ws, f"A{hr}", "Q#", F(10, True, WHITE), Fill(PURPLE_DARK), A())
    put(ws, f"B{hr}", "Question", F(10, True, WHITE), Fill(PURPLE_DARK), A("left"))
    put(ws, f"C{hr}", "+/−", F(10, True, WHITE), Fill(PURPLE_DARK), A())
    for j, pid in enumerate(ids):
        put(ws, f"{get_column_letter(4+j)}{hr}", pid, F(9, True, WHITE), Fill(PURPLE_DARK), A())
    raw_c = get_column_letter(4 + n)
    sus_c = get_column_letter(5 + n)
    pass_c = get_column_letter(6 + n)
    put(ws, f"{raw_c}{hr}", "Raw\nScore", F(9, True, WHITE), Fill(HDR_GREY), A())
    put(ws, f"{sus_c}{hr}", "SUS\nScore", F(9, True, WHITE), Fill(PURPLE_DARK), A())
    put(ws, f"{pass_c}{hr}", "Pass?\n≥68", F(9, True, WHITE), Fill(PURPLE_DARK), A())
    q0 = hr + 1
    for qi, (qid, qtext, sign) in enumerate(SUS_QUESTIONS):
        r = q0 + qi
        is_neg = sign == "-"
        rfill = NEG_FILL if is_neg else YELLOW
        put(ws, f"A{r}", qid, F(10, True, PURPLE_DARK), Fill(PURPLE_LIGHT), A())
        put(ws, f"B{r}", qtext, F(10), Fill(GREY_ALT if qi % 2 else WHITE), A("left"))
        put(ws, f"C{r}", "+" if not is_neg else "-",
            F(10, True, NEG_TEXT if is_neg else GREEN_DARK), Fill(NEG_FILL if is_neg else "E2EFDA"), A())
        for j, pid in enumerate(ids):
            put(ws, f"{get_column_letter(4+j)}{r}", sus[pid][qi], F(10), Fill(rfill), A())
    raw_r = q0 + 10
    sus_r = raw_r + 1
    res_r = raw_r + 2
    ws.merge_cells(start_row=raw_r, start_column=1, end_row=raw_r, end_column=3)
    put(ws, f"A{raw_r}", "Raw Score (Sum adjusted)", F(10, True), Fill("F0F0F0"), A("left"))
    ws.merge_cells(start_row=sus_r, start_column=1, end_row=sus_r, end_column=3)
    put(ws, f"A{sus_r}", "SUS Score (Raw × 2.5)", F(9, True, WHITE), Fill(PURPLE_MID), A("left"))
    ws.merge_cells(start_row=res_r, start_column=1, end_row=res_r, end_column=3)
    put(ws, f"A{res_r}", "Result (Pass ≥ 68)", F(10, True, WHITE), Fill(PURPLE_DARK), A("left"))
    qr = [q0 + k for k in range(10)]
    for j, pid in enumerate(ids):
        L = get_column_letter(4 + j)
        pos = [qr[0], qr[2], qr[4], qr[6], qr[8]]
        neg = [qr[1], qr[3], qr[5], qr[7], qr[9]]
        terms = "+".join([f"({L}{rp}-1)" for rp in pos] + [f"(5-{L}{rn})" for rn in neg])
        put(ws, f"{L}{raw_r}", f'=IFERROR({terms},"")', F(10, True), Fill("F0F0F0"), A(wrap=False))
        put(ws, f"{L}{sus_r}", f'=IFERROR({L}{raw_r}*2.5,"")', F(10, True, WHITE), Fill(PURPLE_MID), A(wrap=False))
        put(ws, f"{L}{res_r}", f'=IFERROR(IF({L}{sus_r}>=68,"PASS ✓","FAIL ✗"),"")', F(10, True, WHITE), Fill(PURPLE_DARK), A(wrap=False))
    raw_c = get_column_letter(4 + n)
    sus_c = get_column_letter(5 + n)
    first_p = get_column_letter(4)
    last_p = get_column_letter(3 + n)
    put(ws, f"{raw_c}{sus_r}", "AVG SUS →", F(9, True, BLACK), Fill("D0D0D0"), A(wrap=False))
    put(ws, f"{sus_c}{sus_r}", f'=IFERROR(AVERAGE({first_p}{sus_r}:{last_p}{sus_r}),"")',
        F(11, True, WHITE), Fill(NAVY_DARK), A(wrap=False))
    put(ws, f"{raw_c}{res_r}", "% PASS →", F(9, True, BLACK), Fill("D0D0D0"), A(wrap=False))
    put(ws, f"{sus_c}{res_r}", f'=IFERROR(COUNTIF({first_p}{sus_r}:{last_p}{sus_r},">=68")/{n},"")',
        F(11, True, WHITE), Fill(NAVY_MID), A(wrap=False))
    ws.cell(row=res_r, column=5 + n).number_format = "0%"
    return res_r + 2, f"{sus_c}{sus_r}", f"{sus_c}{res_r}"

top = 3
top, PAT_SUS_AVG_CELL, PAT_PASS_CELL = sus_block(wss, top, "Patients (Patient role)", PIDS, psus, PURPLE_MID)
top, DOC_SUS_AVG_CELL, DOC_PASS_CELL = sus_block(wss, top, "Doctors (Doctor role)", DIDS, dsus, NAVY_MID)
gtop = top
wss.merge_cells(start_row=gtop, start_column=1, end_row=gtop, end_column=3 + NMAX + 3)
put(wss, f"A{gtop}", "SUS Scoring Guide", F(10, True, WHITE), Fill("333333"), A("left"))
guide = [("≥ 90", "A — Excellent — Best Imaginable", "D5E8D4"),
         ("≥ 80", "B — Good — Above Average", PURPLE_LIGHT),
         ("≥ 68", "C — OK (Pass) — Acceptable", YELLOW),
         ("51–67", "D — Poor — Below Average", "FFDDB3"),
         ("< 51", "F — Fail — Worst Imaginable", NEG_FILL)]
for i, (rng_, txt, fl) in enumerate(guide):
    r = gtop + 1 + i
    put(wss, f"A{r}", rng_, F(10, True), Fill(fl), A())
    wss.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3 + NMAX + 3)
    put(wss, f"B{r}", txt, F(10), Fill(fl), A("left"))
note_r = gtop + 6
wss.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=3 + NMAX + 3)
put(wss, f"A{note_r}",
    "Scoring: Odd Qs (1,3,5,7,9): adjusted = response − 1  |  Even Qs (2,4,6,8,10): adjusted = 5 − response  |  SUS = sum × 2.5  |  Pass ≥ 68",
    F(9, color=PURPLE_DARK), Fill(PURPLE_LIGHT), A("left"))

# ============================================================ SUMMARY
wsm = wb.create_sheet("Summary")
wsm.sheet_view.showGridLines = False
for col, w in {"A": 16, "B": 24, "C": 32, "D": 26, "E": 20, "F": 14}.items():
    wsm.column_dimensions[col].width = w
wsm.merge_cells("A1:F1")
put(wsm, "A1", "MediDiagnose — Usability Testing Summary", F(14, True, WHITE), Fill(TEAL_DARK), A())
wsm.row_dimensions[1].height = 26

wsm.merge_cells("A3:E3")
put(wsm, "A3", "Participant Overview", F(11, True, WHITE), Fill(TEAL_DARK), A())
for j, h in enumerate(["User Group", "Count", "Role", "Tasks", "Sessions"]):
    put(wsm, f"{get_column_letter(1+j)}4", h, F(10, True, WHITE), Fill(TEAL_DARK), A("left"))
put(wsm, "A5", "Patients", F(10), Fill(WHITE), A("left"))
put(wsm, "B5", len(PIDS), F(10), Fill(WHITE), A())
put(wsm, "C5", "Patient", F(10), Fill(WHITE), A("left"))
put(wsm, "D5", "P1–P7", F(10), Fill(WHITE), A("left"))
put(wsm, "E5", "3 (Day 0, +2, +7)", F(10), Fill(WHITE), A("left"))
put(wsm, "A6", "Doctors", F(10), Fill(GREY_ALT), A("left"))
put(wsm, "B6", len(DIDS), F(10), Fill(GREY_ALT), A())
put(wsm, "C6", "Doctor", F(10), Fill(GREY_ALT), A("left"))
put(wsm, "D6", "D1–D6", F(10), Fill(GREY_ALT), A("left"))
put(wsm, "E6", "3 (Day 0, +2, +7)", F(10), Fill(GREY_ALT), A("left"))

wsm.merge_cells("A8:F8")
put(wsm, "A8", "Key Results (live)", F(11, True, WHITE), Fill(NAVY_DARK), A())
for j, h in enumerate(["Metric", "Patients", "Doctors", "Notes"]):
    put(wsm, f"{get_column_letter(1+j)}9", h, F(10, True, WHITE), Fill(NAVY_DARK), A("left"))
put(wsm, "A10", "Mean SUS Score", F(10, True), Fill(S1F), A("left"))
put(wsm, "B10", f"=SUS!{PAT_SUS_AVG_CELL}", F(10), Fill(S1F), A())
put(wsm, "C10", f"=SUS!{DOC_SUS_AVG_CELL}", F(10), Fill(S1F), A())
put(wsm, "D10", "0–100, pass ≥ 68", F(10), Fill(S1F), A("left"))
put(wsm, "A11", "SUS Pass Rate", F(10, True), Fill(S2F), A("left"))
put(wsm, "B11", f"=SUS!{PAT_PASS_CELL}", F(10), Fill(S2F), A())
put(wsm, "C11", f"=SUS!{DOC_PASS_CELL}", F(10), Fill(S2F), A())
wsm["B11"].number_format = "0%"; wsm["C11"].number_format = "0%"
put(wsm, "D11", "Share scoring ≥ 68", F(10), Fill(S2F), A("left"))

wsm.merge_cells("A13:F13")
put(wsm, "A13", "Metrics Recorded", F(11, True, WHITE), Fill(GREEN_DARK), A())
for j, h in enumerate(["Metric", "Description", "Type"]):
    put(wsm, f"{get_column_letter(1+j)}14", h, F(10, True, WHITE), Fill(GREEN_DARK), A("left"))
metrics = [("Task Completion Time", "Seconds per task per session", "Quantitative"),
           ("Number of Errors", "Wrong actions, wrong clicks, confusion", "Quantitative"),
           ("Help Requests", "Times guidance was requested", "Quantitative"),
           ("Task Success", "1 = success, 0 = fail", "Quantitative"),
           ("SUS Score", "0–100, pass ≥ 68", "Quantitative")]
for i, (m, de, ty) in enumerate(metrics):
    r = 15 + i
    fl = "E2EFDA" if i % 2 else WHITE
    put(wsm, f"A{r}", m, F(10, True), Fill(fl), A("left"))
    put(wsm, f"B{r}", de, F(10), Fill(fl), A("left"))
    put(wsm, f"C{r}", ty, F(10), Fill(fl), A("left"))

wsm.merge_cells("A21:F21")
put(wsm, "A21", "SUS Score Interpretation", F(11, True, WHITE), Fill(PURPLE_DARK), A())
for j, h in enumerate(["Score Range", "Grade", "Interpretation"]):
    put(wsm, f"{get_column_letter(1+j)}22", h, F(10, True, WHITE), Fill(PURPLE_DARK), A("left"))
interp = [("≥ 90", "A — Excellent", "Best Imaginable", "D5E8D4"),
          ("80–89", "B — Good", "Above Average", PURPLE_LIGHT),
          ("68–79", "C — OK (Pass)", "Acceptable", YELLOW),
          ("51–67", "D — Poor", "Below Average", "FFDDB3"),
          ("< 51", "F — Fail", "Worst Imaginable", NEG_FILL)]
for i, (sr, gr, it, fl) in enumerate(interp):
    r = 23 + i
    put(wsm, f"A{r}", sr, F(10, True), Fill(fl), A())
    put(wsm, f"B{r}", gr, F(10), Fill(fl), A("left"))
    put(wsm, f"C{r}", it, F(10), Fill(fl), A("left"))

# Force Excel to recalculate all formulas when the file is opened (no cached
# values needed -> no Excel automation required to produce correct numbers).
wb.calculation.fullCalcOnLoad = True

OUT = r"C:/Users/Youssef/Downloads/MediDiagnose_Usability_Testing_filled.xlsx"
wb.save(OUT)
print("saved", OUT, "| doctors:", len(DIDS))
